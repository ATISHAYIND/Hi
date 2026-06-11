#!/usr/bin/env python3
"""
Multi-account EBS Snapshot AUDIT  ->  management-ready Excel report
------------------------------------------------------------------
Read-only. NEVER deletes anything.

Walks every logged-in AWS SSO profile, scans all regions, finds EBS
snapshots older than N days (default 90), and writes a formatted .xlsx
with an executive dashboard + a fully self-explanatory detail sheet
(snapshot id, age, plain-English reason, recommended action, etc.).

Excluded from the "deletable" view (flagged, never touched):
  - tagged 'aws:backup:source-resource'  -> AWS Backup managed
  - in use by a registered AMI

Usage:
  pip install boto3 openpyxl
  python snapshot_audit.py
  python snapshot_audit.py --days 120
  python snapshot_audit.py --regions ap-south-1 us-east-1
  python snapshot_audit.py --out Snapshot_Audit_Report.xlsx
"""

import argparse
import subprocess
import sys
from datetime import datetime, timezone, timedelta

import boto3
from botocore.exceptions import ClientError, BotoCoreError

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------- config -----------------------------
BACKUP_TAG_KEY = "aws:backup:source-resource"
GB_MONTH_COST = 0.05            # approx USD/GB-month, standard EBS snapshot (ap-south-1)
DEFAULT_STS_REGION = "ap-south-1"
# ------------------------------------------------------------------

# ------------------------------ palette ---------------------------
NAVY   = "1F3864"
BLUE   = "2F5496"
HEADER = "305496"
WHITE  = "FFFFFF"
CARD   = "EAF1FB"
BAND   = "F4F8FD"
GREEN_F, GREEN_T = "C6EFCE", "006100"
AMBER_F, AMBER_T = "FFEB9C", "9C6500"
GREY_F,  GREY_T  = "DDDDDD", "3F3F3F"
LINE = "BFBFBF"
FONT = "Calibri"
# ------------------------------------------------------------------


def get_all_profiles():
    try:
        profiles = subprocess.check_output(
            ["aws", "configure", "list-profiles"], text=True
        ).splitlines()
        if not profiles:
            raise Exception("No AWS profiles found. Please configure at least one profile.")
        valid = []
        for p in profiles:
            try:
                boto3.Session(profile_name=p).client(
                    "sts", region_name=DEFAULT_STS_REGION).get_caller_identity()
                valid.append(p)
            except Exception:
                print(f"  [!] Profile '{p}' not logged in or token expired. Skipping...")
        if not valid:
            raise Exception("No AWS SSO profiles are currently logged in. "
                            "Run 'aws sso login --profile <profile>'.")
        return valid
    except Exception as e:
        print(f"Error detecting AWS profiles: {e}")
        sys.exit(1)


def get_regions(session, override):
    if override:
        return override
    ec2 = session.client("ec2", region_name=DEFAULT_STS_REGION)
    return [r["RegionName"] for r in ec2.describe_regions()["Regions"]]


def get_account_id(session):
    return session.client("sts", region_name=DEFAULT_STS_REGION).get_caller_identity()["Account"]


def has_backup_tag(s):
    return any(t["Key"] == BACKUP_TAG_KEY for t in s.get("Tags", []))


def name_tag(s):
    for t in s.get("Tags", []):
        if t["Key"] == "Name":
            return t["Value"]
    return ""


def ami_snapshot_ids(ec2):
    in_use = set()
    try:
        for img in ec2.describe_images(Owners=["self"])["Images"]:
            for bdm in img.get("BlockDeviceMappings", []):
                snap = bdm.get("Ebs", {}).get("SnapshotId")
                if snap:
                    in_use.add(snap)
    except (ClientError, BotoCoreError) as e:
        print(f"      [!] Could not list AMIs in region: {e}")
    return in_use


def scan_region(session, account_id, region, cutoff):
    ec2 = session.client("ec2", region_name=region)
    in_use = ami_snapshot_ids(ec2)
    rows = []
    for page in ec2.get_paginator("describe_snapshots").paginate(OwnerIds=["self"]):
        for s in page["Snapshots"]:
            start = s["StartTime"]
            if start >= cutoff:
                continue
            sid = s["SnapshotId"]
            if has_backup_tag(s):
                status = "SKIP_BACKUP"
            elif sid in in_use:
                status = "SKIP_AMI"
            else:
                status = "DELETABLE"
            size = s.get("VolumeSize", 0)
            rows.append({
                "account_id": account_id, "region": region, "snapshot_id": sid,
                "volume_id": s.get("VolumeId", "vol-?"),
                "name": name_tag(s),
                "description": s.get("Description", "") or "",
                "size_gb": size,
                "age_days": (datetime.now(timezone.utc) - start).days,
                "start_time": start.strftime("%Y-%m-%d"),
                "encrypted": "Yes" if s.get("Encrypted") else "No",
                "storage_tier": (s.get("StorageTier") or "standard").title(),
                "status": status,
                "est_monthly_usd": round(size * GB_MONTH_COST, 2),
            })
    return rows


def reason_action(status, age):
    if status == "DELETABLE":
        return (f"{age} days old. No AWS Backup tag and not used by any AMI "
                f"- eligible for deletion.", "Review & delete")
    if status == "SKIP_BACKUP":
        return (f"{age} days old. Managed by AWS Backup (aws:backup:source-resource "
                f"tag) - do NOT delete manually.", "Retain - AWS Backup managed")
    return (f"{age} days old. Backs a registered AMI - deregister the AMI before "
            f"any deletion.", "Retain - AMI in use")


# ---------------------------- excel build -------------------------
def _thin(color=LINE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _kpi_box(ws, col, row, label, value, fill=CARD, vcolor=NAVY, big=20):
    c1, c2 = get_column_letter(col), get_column_letter(col + 1)
    ws.merge_cells(f"{c1}{row}:{c2}{row}")
    ws.merge_cells(f"{c1}{row+1}:{c2}{row+1}")
    lab = ws[f"{c1}{row}"]
    lab.value = label
    lab.font = Font(name=FONT, size=9, bold=True, color="555555")
    lab.alignment = Alignment(horizontal="center", vertical="center")
    val = ws[f"{c1}{row+1}"]
    val.value = value
    val.font = Font(name=FONT, size=big, bold=True, color=vcolor)
    val.alignment = Alignment(horizontal="center", vertical="center")
    for r in (row, row + 1):
        for cc in (col, col + 1):
            cell = ws.cell(row=r, column=cc)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = _thin()
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 30
    return val


def build_workbook(rows, meta, out_path):
    wb = Workbook()

    # ---------- Details sheet ----------
    d = wb.active
    d.title = "Snapshot Details"
    headers = ["Account ID", "Region", "Snapshot ID", "Volume ID", "Name",
               "Description", "Size (GB)", "Age (days)", "Created", "Encrypted",
               "Storage Tier", "Status", "Reason / Justification",
               "Recommended Action", "Est. Monthly ($)"]
    NCOL = len(headers)                      # 15
    SIZE_C, STATUS_C, COST_C, ACC_C = 7, 12, 15, 1
    d.append(headers)
    for c in range(1, NCOL + 1):
        cell = d.cell(row=1, column=c)
        cell.font = Font(name=FONT, size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()
    d.row_dimensions[1].height = 30

    status_style = {
        "DELETABLE":   (GREEN_F, GREEN_T),
        "SKIP_BACKUP": (AMBER_F, AMBER_T),
        "SKIP_AMI":    (GREY_F,  GREY_T),
    }
    left_cols = {5, 6, 13}                    # Name, Description, Reason -> left aligned
    for i, r in enumerate(rows, start=2):
        reason, action = reason_action(r["status"], r["age_days"])
        d.append([r["account_id"], r["region"], r["snapshot_id"], r["volume_id"],
                  r["name"], r["description"], r["size_gb"], r["age_days"],
                  r["start_time"], r["encrypted"], r["storage_tier"],
                  r["status"], reason, action, r["est_monthly_usd"]])
        band = BAND if i % 2 == 0 else WHITE
        for c in range(1, NCOL + 1):
            cell = d.cell(row=i, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.fill = PatternFill("solid", fgColor=band)
            cell.border = _thin("E0E0E0")
            cell.alignment = Alignment(
                horizontal="left" if c in left_cols else "center",
                vertical="center",
                wrap_text=(c in {6, 13}))
        d.cell(row=i, column=SIZE_C).number_format = "#,##0"
        d.cell(row=i, column=COST_C).number_format = '$#,##0.00'
        f, t = status_style[r["status"]]
        sc = d.cell(row=i, column=STATUS_C)
        sc.fill = PatternFill("solid", fgColor=f)
        sc.font = Font(name=FONT, size=10, bold=True, color=t)
        d.row_dimensions[i].height = 30

    last = len(rows) + 1
    tr = last + 1
    d.cell(row=tr, column=6, value="TOTAL").font = Font(name=FONT, bold=True)
    d.cell(row=tr, column=6).alignment = Alignment(horizontal="right")
    d.cell(row=tr, column=SIZE_C, value=f"=SUM({get_column_letter(SIZE_C)}2:{get_column_letter(SIZE_C)}{last})")
    d.cell(row=tr, column=COST_C, value=f"=SUM({get_column_letter(COST_C)}2:{get_column_letter(COST_C)}{last})")
    for c in (SIZE_C, COST_C):
        cell = d.cell(row=tr, column=c)
        cell.font = Font(name=FONT, bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=CARD)
    d.cell(row=tr, column=SIZE_C).number_format = "#,##0"
    d.cell(row=tr, column=COST_C).number_format = '$#,##0.00'

    widths = [14, 12, 23, 23, 18, 26, 9, 9, 12, 10, 11, 14, 46, 22, 14]
    for i, w in enumerate(widths, start=1):
        d.column_dimensions[get_column_letter(i)].width = w
    d.freeze_panes = "C2"
    d.auto_filter.ref = f"A1:{get_column_letter(NCOL)}{last}"
    d.sheet_view.showGridLines = False

    # ---------- Executive Summary ----------
    s = wb.create_sheet("Executive Summary", 0)
    s.sheet_view.showGridLines = False
    for i, w in enumerate([3, 16, 16, 16, 16, 16, 16, 16, 16, 3], start=1):
        s.column_dimensions[get_column_letter(i)].width = w

    s.merge_cells("B2:I2")
    t = s["B2"]
    t.value = "EBS Snapshot Cost-Optimization  -  Audit Report"
    t.font = Font(name=FONT, size=18, bold=True, color=WHITE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, 10):
        s.cell(row=2, column=c).fill = PatternFill("solid", fgColor=NAVY)
    s.row_dimensions[2].height = 38

    s.merge_cells("B3:I3")
    sub = s["B3"]
    sub.value = (f"Generated {meta['generated']}   |   Snapshots older than "
                 f"{meta['days']} days   |   Accounts scanned: {meta['accounts']}"
                 f"   |   Read-only audit")
    sub.font = Font(name=FONT, size=10, italic=True, color="FFFFFF")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, 10):
        s.cell(row=3, column=c).fill = PatternFill("solid", fgColor=BLUE)
    s.row_dimensions[3].height = 20

    DET = "'Snapshot Details'"
    SZ, ST, CO, AC = (get_column_letter(SIZE_C), get_column_letter(STATUS_C),
                      get_column_letter(COST_C), get_column_letter(ACC_C))
    _kpi_box(s, 2, 5, "OLD SNAPSHOTS (TOTAL)", len(rows), CARD, NAVY)
    _kpi_box(s, 4, 5, "DELETABLE",
             f'=COUNTIF({DET}!{ST}:{ST},"DELETABLE")', GREEN_F, GREEN_T)
    _kpi_box(s, 6, 5, "EXCLUDED - AWS BACKUP",
             f'=COUNTIF({DET}!{ST}:{ST},"SKIP_BACKUP")', AMBER_F, AMBER_T)
    _kpi_box(s, 8, 5, "EXCLUDED - AMI-BACKED",
             f'=COUNTIF({DET}!{ST}:{ST},"SKIP_AMI")', GREY_F, GREY_T)

    _kpi_box(s, 2, 8, "DELETABLE STORAGE (GB)",
             f'=SUMIF({DET}!{ST}:{ST},"DELETABLE",{DET}!{SZ}:{SZ})', CARD, NAVY)
    mb = _kpi_box(s, 4, 8, "EST. MONTHLY SAVING",
                  f'=SUMIF({DET}!{ST}:{ST},"DELETABLE",{DET}!{CO}:{CO})', "E2EFDA", GREEN_T)
    mb.number_format = '$#,##0'
    ab = _kpi_box(s, 6, 8, "EST. ANNUAL SAVING",
                  f'=SUMIF({DET}!{ST}:{ST},"DELETABLE",{DET}!{CO}:{CO})*12', "E2EFDA", GREEN_T)
    ab.number_format = '$#,##0'
    gbtot = _kpi_box(s, 8, 8, "TOTAL OLD STORAGE (GB)",
                     f'=SUM({DET}!{SZ}2:{SZ}{last})', CARD, NAVY)
    gbtot.number_format = '#,##0'

    hr = 12
    s.merge_cells(f"B{hr}:I{hr}")
    bh = s[f"B{hr}"]
    bh.value = "Breakdown by Account"
    bh.font = Font(name=FONT, size=12, bold=True, color=NAVY)
    hr += 1
    cols = ["Account ID", "Total Old", "Deletable", "Excl. Backup",
            "Excl. AMI", "Deletable GB", "Est. Monthly ($)"]
    for j, h in enumerate(cols, start=2):
        cell = s.cell(row=hr, column=j, value=h)
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin()
    accounts = sorted({r["account_id"] for r in rows})
    rr = hr + 1
    for k, acc in enumerate(accounts):
        a = f'{DET}!{AC}:{AC}'
        s.cell(row=rr, column=2, value=acc)
        s.cell(row=rr, column=3, value=f'=COUNTIF({a},$B{rr})')
        s.cell(row=rr, column=4,
               value=f'=COUNTIFS({a},$B{rr},{DET}!{ST}:{ST},"DELETABLE")')
        s.cell(row=rr, column=5,
               value=f'=COUNTIFS({a},$B{rr},{DET}!{ST}:{ST},"SKIP_BACKUP")')
        s.cell(row=rr, column=6,
               value=f'=COUNTIFS({a},$B{rr},{DET}!{ST}:{ST},"SKIP_AMI")')
        s.cell(row=rr, column=7,
               value=f'=SUMIFS({DET}!{SZ}:{SZ},{a},$B{rr},{DET}!{ST}:{ST},"DELETABLE")')
        s.cell(row=rr, column=8,
               value=f'=SUMIFS({DET}!{CO}:{CO},{a},$B{rr},{DET}!{ST}:{ST},"DELETABLE")')
        band = BAND if k % 2 == 0 else WHITE
        for j in range(2, 9):
            cell = s.cell(row=rr, column=j)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor=band)
            cell.border = _thin("E0E0E0")
        s.cell(row=rr, column=7).number_format = "#,##0"
        s.cell(row=rr, column=8).number_format = '$#,##0.00'
        rr += 1

    rr += 1
    s.merge_cells(f"B{rr}:I{rr+2}")
    note = s[f"B{rr}"]
    note.value = ("Methodology: Self-owned snapshots older than the threshold across all "
                  "enabled regions of every logged-in account. Snapshots tagged "
                  "'aws:backup:source-resource' (AWS Backup managed) and snapshots backing a "
                  "registered AMI are EXCLUDED from deletable totals. Savings are an upper "
                  "bound - EBS snapshots are incremental, so realised savings will be lower. "
                  "This report is read-only; no resources were modified.")
    note.font = Font(name=FONT, size=9, italic=True, color="555555")
    note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="Read-only old EBS snapshot audit -> Excel")
    ap.add_argument("--days", type=int, default=90)
    ap.add_argument("--regions", nargs="*", default=None)
    ap.add_argument("--out", default="Snapshot_Audit_Report.xlsx")
    args = ap.parse_args()

    cutoff = datetime.now(timezone.utc) - timedelta(days=args.days)
    print("\n=== Detecting logged-in SSO profiles ===")
    profiles = get_all_profiles()
    print(f"  -> {len(profiles)} active profile(s): {', '.join(profiles)}\n")

    rows = []
    for p in profiles:
        session = boto3.Session(profile_name=p)
        try:
            acc = get_account_id(session)
        except Exception as e:
            print(f"[{p}] could not resolve account id: {e}")
            continue
        print(f"=== Account {acc}  (profile: {p}) ===")
        for region in get_regions(session, args.regions):
            try:
                rr = scan_region(session, acc, region, cutoff)
            except (ClientError, BotoCoreError) as e:
                print(f"  [{region}] skipped: {e}")
                continue
            rows.extend(rr)
            if rr:
                print(f"  [{region}] found {len(rr)} old snapshot(s)")

    meta = {
        "generated": datetime.now().strftime("%d %b %Y, %H:%M"),
        "days": args.days,
        "accounts": len({r["account_id"] for r in rows}) or len(profiles),
    }
    build_workbook(rows, meta, args.out)
    print(f"\nDone. {len(rows)} snapshots written to: {args.out}")


if __name__ == "__main__":
    main()
